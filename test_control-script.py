import pandas as pd
import numpy as np
import itertools
import pytest
import ipdb

"""
NOTES
dcms..._2016_tables_2.xlsx is the workbook we are checking against and adds some fixes: 
    adds some missing anonymisations. below is a non exhaustive list
        I previously added some missing anonymisation to some tables (sex I think?) but didn't make a note of it - think I emailed penny though
        add 100 to all regions perc on region table to make consistent with other region tables.
        add anonymisation to digital sector by region (north east)
        add anonymisation to sport sector by region (east midlands)
    
    rounds qualification and ftpt (rounded .5 down in some cases) numbers to 0dp

however, this approach does not make sense to add in over-anonymisation (since I don't necessarily know what the correct value should be), instead temporarily remove these cases from my data. make a list below:
    sex
        employed
            total
                telecoms
    civil society region northern ireland total
                

we are currently annonymising overlap row before it is used for all_dcms which needs fixing

want to keep all cat specific adjustments in this main script, not in functions
"""


current_year = 2016

allyears = {}
for year in range(2012, current_year + 1):
    allyears[year] = pd.read_csv("~/data/cleaned_" + str(year) + "_df.csv")
    #print("~/data/cleaned_" + str(year) + "_df.csv")


# df2016 = pd.read_csv("~/data/cleaned_2016_df.csv")
#df2016 = allyears[current_year]

# finish cleaning data post R cleaning ========================================

regionlookupdata = pd.read_csv('~/projects/employment/region-lookup.csv')
regionlookdict = {}
for index, row in regionlookupdata.iterrows():
    regionlookdict.update({row[0] : row[1]})


def clean_raw_data(df):
        
    df['regionmain'] = df.GORWKR.map(regionlookdict)
    df['regionsecond'] = df.GORWK2R.map(regionlookdict)
    
    return df

allyears = {k: clean_raw_data(v) for k, v in allyears.items()}

df = allyears[current_year]    
df['qualification'] = df['qualification'].astype(str)
df['ftpt'] = df['ftpt'].astype(str)
df['nssec'] = df['nssec'].astype(str)

sic_mappings = pd.read_csv("~/projects/employment/sic_mappings.csv")
sic_mappings = sic_mappings[sic_mappings.sic != 62.011]
sic_mappings.sic = round(sic_mappings.sic * 100, 0)
sic_mappings.sic = sic_mappings.sic.astype(int)


from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# write data to excel template ================================================
    
wb = load_workbook('DCMS_Sectors_Economic_Estimates_Employment_2016_tables_Template_unmerged.xlsx')

"""
#catvar <- "sex"; catorder <- c("Male", "Female"); sheet <- 14; xy <- c(2,9); perc <- TRUE; cattotal <- TRUE
catvar <- "ethnicity"; catorder <- c("White", "BAME"); sheet <- 15; xy <- c(2,8); perc <- TRUE; cattotal <- TRUE
#catvar <- "dcms_ageband"; catorder <- NA; sheet <- 16; xy <- c(2,8); perc <- FALSE; cattotal <- TRUE
#catvar <- "qualification"; catorder <- NA; sheet <- 17; xy <- c(2,7); perc <- FALSE; catorder <- c("Degree or equivalent",	"Higher Education",	"A Level or equivalent", "GCSE A* - C or equivalent",	"Other",	"No Qualification"); cattotal <- TRUE
#catvar <- "ftpt"; catorder <- c("Full time", "Part time"); sheet <- 18; xy <- c(2,8); perc <- TRUE; cattotal <- TRUE
#catvar <- "nssec"; catorder <- c("More Advantaged Group (NS-SEC 1-4)", "Less Advantaged Group (NS-SEC 5-8)"); sheet <- 19; xy <- c(2,8); perc <- FALSE; cattotal <- FALSE


if (catvar == "qualification") df <- df[df[, catvar] != "dont know" & !is.na(df[, catvar]), ]
if (catvar == "ftpt") df <- df[df[, catvar] %in% catorder, ]
if (catvar == "nssec") df <- df[df[, catvar] %in% catorder, ]
"""


# make time series tables =====================================================

# iterate over each years data to make time series
data = []
for k, v in allyears.items():
    #d2[k] = f(v)
    # CLEANING DATA - adding up main and second jobs, calculating some totals, columns for sector, cat, region, count
    # there doesn't appear to be any tables which use both region and a demographic category, so simply remove region or replace cat column with it.
    sic_level = False
    if k < 2016:
        weightedcountcol = 'PWTA14'
    if k == 2016:
        weightedcountcol = 'PWTA16'
    if k == 2017:
        weightedcountcol = 'PWTA17'
    
    import clean_data_func
    allyears[k] = clean_data_func.clean_data(v, {'mycat': 'region'}, sic_mappings, regionlookupdata, weightedcountcol)
    
    import aggregate_data_func
    allyears[k] = aggregate_data_func.aggregate_data(allyears[k], {'mycat': 'region'})
    
    mycol = allyears[k][allyears[k]['emptype'] == 'total']
    mycol = mycol.groupby(['sector']).sum()
    mycol = mycol.rename(columns={'count': k})
    data.append(mycol)

timeseries = pd.concat(data, axis=1)

if current_year < 2016:
    weightedcountcol = 'PWTA14'
if current_year == 2016:
    weightedcountcol = 'PWTA16'
if current_year == 2017:
    weightedcountcol = 'PWTA17'


# iterate through different cats to make cat tables for current year
differencelist = {}
for table in ['sex', 'time_series', 'ethnicity', 'dcms_ageband', 'qualification', 'ftpt', 'nssec', 'region', 'cs', 'ci', 'culture', 'digital', 'gambling', 'sport', 'telecoms']:
    
    anoncats = []
    mycat = None
    perc = None
    cattotal = None
    catorder = []
    emptypecats = None
    wsname = None
    startrow = None
    startcol = None
    finishrow = None
    finishcol = None
    region = None
    sector = np.nan
    time_series = None
    
    # set default param dict
    table_params = {
        'anoncats': [],
        'mycat': None,
        'perc': None,
        'cattotal': None,
        'catorder': [],
        'emptypecats': None,
        'wsname': None,
        'startrow': None,
        'startcol': None,
        'finishrow': None,
        'finishcol': None,
        'region': None,
        'sector': np.nan,
        'time_series': None
    }
        
    # sex
    if table == 'sex':
        mycat = 'sex'
        perc = True
        cattotal = True
        catorder = ['Male', 'Female']
        emptypecats = True
        wsname = "3.5 - Gender (000's)"
        startrow = 9
        startcol = 2
        finishrow = 17
        finishcol = 16
        table_params.update({
            'mycat': 'sex',
            'perc': True,
            'cattotal': True,
            'catorder': ['Male', 'Female'],
            'emptypecats': True,
            'wsname': "3.5 - Gender (000's)",
            'startrow': 9,
            'startcol': 2,
            'finishrow': 17,
            'finishcol': 16
        })
        
    # ethnicity
    if table == 'ethnicity':
        mycat = 'ethnicity'
        perc = True
        cattotal = True
        catorder = ['White', 'BAME']
        emptypecats = False
        wsname = "3.6 - Ethnicity (000's)"
        startrow = 8
        startcol = 2
        finishrow = 16
        finishcol = 6
        table_params.update({
            'mycat': 'ethnicity',
            'perc': True,
            'cattotal': True,
            'catorder': ['White', 'BAME'],
            'emptypecats': False,
            'wsname': "3.6 - Ethnicity (000's)",
            'startrow': 8,
            'startcol': 2,
            'finishrow': 16,
            'finishcol': 6
        })
        
        # ethnicity
    if table == 'dcms_ageband':
        mycat = 'dcms_ageband'
        perc = False
        cattotal = True
        catorder = ['16-24 years', '25-39 years', '40-59 years', '60 years +']
        emptypecats = True
        wsname = "3.7 - Age (000's)"
        startrow = 8
        startcol = 2
        finishrow = 16
        finishcol = 16
        table_params.update({
            'mycat': 'dcms_ageband',
            'perc': False,
            'cattotal': True,
            'catorder': ['16-24 years', '25-39 years', '40-59 years', '60 years +'],
            'emptypecats': True,
            'wsname': "3.7 - Age (000's)",
            'startrow': 8,
            'startcol': 2,
            'finishrow': 16,
            'finishcol': 16
        })
        
    if table == 'qualification':
        mycat = 'qualification'
        perc = False
        cattotal = True
        catorder = ["Degree or equivalent",	"Higher Education",	"A Level or equivalent", "GCSE A* - C or equivalent",	"Other",	"No Qualification"]
        emptypecats = False
        wsname = "3.8 - Qualification (000's)"
        startrow = 7
        startcol = 2
        finishrow = 15
        finishcol = 8
        anoncats = ['Other', 'No Qualification']
        table_params.update({
            'mycat': 'qualification',
            'perc': False,
            'cattotal': True,
            'catorder': ["Degree or equivalent",	"Higher Education",	"A Level or equivalent", "GCSE A* - C or equivalent",	"Other",	"No Qualification"],
            'emptypecats': False,
            'wsname': "3.8 - Qualification (000's)",
            'startrow': 7,
            'startcol': 2,
            'finishrow': 15,
            'finishcol': 8,
            'anoncats': ['Other', 'No Qualification']
        })
                
    if table == 'ftpt':
        mycat = 'ftpt'
        perc = True
        cattotal = True
        catorder = ['Full time', 'Part time']
        emptypecats = False
        wsname = "3.9 - Fulltime Parttime (000's)"
        startrow = 8
        startcol = 2
        finishrow = 16
        finishcol = 6
        table_params.update({
            'mycat': 'ftpt',
            'perc': True,
            'cattotal': True,
            'catorder': ['Full time', 'Part time'],
            'emptypecats': False,
            'wsname': "3.9 - Fulltime Parttime (000's)",
            'startrow': 8,
            'startcol': 2,
            'finishrow': 16,
            'finishcol': 6
        })

    if table == 'nssec':
        mycat = 'nssec'
        perc = False
        cattotal = False
        catorder = ["More Advantaged Group (NS-SEC 1-4)", "Less Advantaged Group (NS-SEC 5-8)"]
        emptypecats = True
        wsname = "3.10 - NS-SEC (000's)"
        startrow = 8
        startcol = 2
        finishrow = 16
        finishcol = 7
        table_params.update({
            'mycat': 'nssec',
            'perc': False,
            'cattotal': False,
            'catorder': ["More Advantaged Group (NS-SEC 1-4)", "Less Advantaged Group (NS-SEC 5-8)"],
            'emptypecats': True,
            'wsname': "3.10 - NS-SEC (000's)",
            'startrow': 8,
            'startcol': 2,
            'finishrow': 16,
            'finishcol': 6
        })
                
    if table == 'region':
        mycat = 'region'
        perc = True
        cattotal = False
        emptypecats = True
        wsname = "3.3 - Region (000's)"
        startrow = 8
        startcol = 2
        finishrow = 21
        finishcol = 8
        table_params.update({
            'mycat': 'region',
            'perc': True,
            'cattotal': False,
            'emptypecats': True,
            'wsname': "3.3 - Region (000's)",
            'startrow': 8,
            'startcol': 2,
            'finishrow': 21,
            'finishcol': 8
        })
        
    if table == 'cs':
        mycat = 'region'
        perc = True
        cattotal = False
        emptypecats = True
        wsname = "3.3a - Civil Society"
        startrow = 8
        startcol = 2
        finishrow = 21
        finishcol = 7
        #region = True
        sector = 'civil_society'
        table_params.update({
            'mycat': 'region',
            'perc': True,
            'cattotal': False,
            'emptypecats': True,
            'wsname': "3.3a - Civil Society",
            'startrow': 8,
            'startcol': 2,
            'finishrow': 21,
            'finishcol': 8,
            'sector': 'civil_society'
        })
        
    if table == 'ci':
        mycat = 'region'
        perc = True
        cattotal = False
        emptypecats = True
        wsname = "3.3b - Creative Industries"
        startrow = 8
        startcol = 2
        finishrow = 21
        finishcol = 7
        region = True
        sector = 'creative'
        table_params.update({
            'mycat': 'region',
            'perc': True,
            'cattotal': False,
            'emptypecats': True,
            'wsname': "3.3a - Civil Society",
            'startrow': 8,
            'startcol': 2,
            'finishrow': 21,
            'finishcol': 8,
            'sector': 'creative'
        })
        
    if table == 'culture':
        mycat = 'region'
        perc = True
        cattotal = False
        emptypecats = True
        wsname = "3.3c - Cultural Sector"
        startrow = 8
        startcol = 2
        finishrow = 21
        finishcol = 7
        region = True
        sector = 'culture'
        table_params.update({
            'mycat': 'region',
            'perc': True,
            'cattotal': False,
            'emptypecats': True,
            'wsname': "3.3c - Cultural Sector",
            'startrow': 8,
            'startcol': 2,
            'finishrow': 21,
            'finishcol': 7,
            'sector': 'culture'
        })
        
    if table == 'digital':
        mycat = 'region'
        perc = True
        cattotal = False
        emptypecats = True
        wsname = "3.3d - Digital Sector"
        startrow = 8
        startcol = 2
        finishrow = 21
        finishcol = 7
        region = True
        sector = 'digital'
        table_params.update({
            'mycat': 'region',
            'perc': True,
            'cattotal': False,
            'emptypecats': True,
            'wsname': "3.3d - Digital Sector",
            'startrow': 8,
            'startcol': 2,
            'finishrow': 21,
            'finishcol': 7,
            'sector': 'digital'
        })
        
    if table == 'gambling':
        mycat = 'region'
        perc = True
        cattotal = False
        emptypecats = True
        wsname = "3.3e - Gambling"
        startrow = 7
        startcol = 2
        finishrow = 20
        finishcol = 3
        region = True
        sector = 'gambling'
        table_params.update({
            'mycat': 'region',
            'perc': True,
            'cattotal': False,
            'emptypecats': True,
            'wsname': "3.3e - Gambling",
            'startrow': 7,
            'startcol': 2,
            'finishrow': 20,
            'finishcol': 3,
            'sector': 'gambling'
        })
        
    if table == 'sport':
        mycat = 'region'
        perc = True
        cattotal = False
        emptypecats = True
        wsname = "3.3f - Sport"
        startrow = 8
        startcol = 2
        finishrow = 21
        finishcol = 7
        region = True
        sector = 'sport'
        table_params.update({
            'mycat': 'region',
            'perc': True,
            'cattotal': False,
            'emptypecats': True,
            'wsname': "3.3f - Sport",
            'startrow': 8,
            'startcol': 2,
            'finishrow': 21,
            'finishcol': 7,
            'sector': 'sport'
        })

    if table == 'telecoms':
        mycat = 'region'
        perc = True
        cattotal = False
        emptypecats = True
        wsname = "3.3g - Telecoms"
        startrow = 7
        startcol = 2
        finishrow = 20
        finishcol = 3
        region = True
        sector = 'telecoms'
        table_params.update({
            'mycat': 'region',
            'perc': True,
            'cattotal': False,
            'emptypecats': True,
            'wsname': "3.3g - Telecoms",
            'startrow': 7,
            'startcol': 2,
            'finishrow': 20,
            'finishcol': 3,
            'sector': 'telecoms'
        })
    
    if table == 'time_series':
        time_series = True
        startrow = 7
        startcol = 3
        finishrow = 17
        finishcol = 7
        wsname = "3.1 - Employment (000's)"
        table_params.update({
            'time_series': True,
            'startrow': 7,
            'startcol': 3,
            'finishrow': 17,
            'finishcol': 7,
            'wsname': "3.1 - Employment (000's)"
        })

    dfcopy = df.copy()
    if mycat == 'qualification':
        dfcopy = dfcopy[dfcopy.qualification != 'dont know']
        dfcopy = dfcopy[dfcopy.qualification != 'nan']
    
    if time_series:
        data = timeseries

        tourism = pd.DataFrame(columns=data.columns)
        tourism.loc['tourism'] = [2, 3, 4, 5, 6]

        percuk = pd.DataFrame(columns=data.columns)
        percuk.loc['percuk'] = [2, 3, 4, 5, 6]
        
        # add tourism
        data = data.append(tourism)
        data = data.append(percuk)
        
        # rounding
        data = round(data / 1000, 0).astype(int)
        
        
        # reorder rows
        myroworder = ["civil_society", "creative", "culture", "digital", "gambling", "sport", "telecoms", 'tourism', "all_dcms", 'percuk', "total_uk"]
        data = data.reindex(myroworder)
        
        
        # CHECK DATA MATCHES PUBLICATION
        # store anonymised values as 0s for comparison and data types
        import check_data_func
        from openpyxl import load_workbook, Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        exceldataframe = check_data_func.check_data(data, wsname, startrow, startcol, finishrow, finishcol)
        # compare computed and publication data
        
        difference = data - exceldataframe
        
        if sum((difference > 1).any()) != 0:
            print(table + ': datasets dont match')
            
    # MAIN GROUP OF FUNCTIONS
    else:        
        cat = mycat
        # CLEANING DATA - adding up main and second jobs, calculating some totals, columns for sector, cat, region, count
        # there doesn't appear to be any tables which use both region and a demographic category, so simply remove region or replace cat column with it.
        sic_level = False
        
        import clean_data_func
        agg = clean_data_func.clean_data(dfcopy, table_params, sic_mappings, regionlookupdata, weightedcountcol)
        
        spsslist = """
        1820, 2611, 2612, 2620, 2630, 2640, 2680, 3012, 3212, 3220, 3230, 4651, 4652, 4763, 4764, 4910, 4932, 4939, 5010, 5030, 5110, 5510, 5520, 5530, 5590, 5610, 5621, 5629, 5630, 5811, 5812, 5813, 5814, 
        5819, 5821, 5829, 5911, 5912, 5913, 5914, 5920, 6010, 6020, 6110, 6120, 6130, 6190, 6201, 6202, 6203, 6209, 6311, 6312, 6391, 6399, 6820, 7021, 7111, 7311, 7312, 7410, 7420, 7430, 7711, 7721, 
        7722, 7729, 7734, 7735, 7740, 7911, 7912, 7990, 8230, 8551, 8552, 9001, 9002, 9003, 9004, 9101, 9102, 9103, 9104, 9200, 9311, 9312, 9313, 9319, 9321, 9329, 9511, 9512 """
        spsslist = spsslist.replace('\n', '')
        spsslist = spsslist.replace('\t', '')
        spsslist = spsslist.replace(' ', '')
        mylist = np.array(spsslist.split(","))
        
        import aggregate_data_func
        aggfinal = aggregate_data_func.aggregate_data(agg, table_params)
            
        if emptypecats == False:
            aggfinal = aggfinal[aggfinal['emptype'] == 'total']
                            
        # SUMMARISING DATA
        if cat == 'region':
            import region_summary_table_func
            final = region_summary_table_func.region_summary_table(aggfinal, table_params)
        else:
            import summary_table_func
            final = summary_table_func.summary_table(aggfinal, cat, perc, cattotal, catorder) 
        
        # ANONYMISING DATA
        import anonymise_func
        data = anonymise_func.anonymise(final, emptypecats, anoncats, cat, sector)
        
        # add extra anonymisation to match publication
        if cat == 'sex':
            data.loc['telecoms', ('employed', 'Total')] = 0
            data.loc['telecoms', ('self employed', 'Total')] = 0
    
        if table == 'cs':
            data.loc['Northern Ireland', 'total'] = 0
            data.loc['Northern Ireland', 'perc_of_all_regions'] = 0
        
        # CHECK DATA MATCHES PUBLICATION
        # store anonymised values as 0s for comparison and data types
        import check_data_func
        from openpyxl import load_workbook, Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        exceldataframe = check_data_func.check_data(data, wsname, startrow, startcol, finishrow, finishcol)
        # compare computed and publication data
        
        difference = data - exceldataframe
        
        if sum((difference > 1).any()) != 0:
            print(cat + ': datasets dont match')
        
        #mylist = make_cat_data()
        #data = mylist[1]
        #difference = mylist[0]

    differencelist.update({table : difference})
    
    ws = wb[wsname]
    rows = dataframe_to_rows(data, index=False, header=False)
    
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
             ws.cell(row=r_idx + startrow - 1, column=c_idx + 1, value=value)
 
  
"""
wsname = "3.5 - Gender (000's)"
startrow = 9
finishrow = 17
finishcol = 16
"""

# marks=pytest.mark.xfail
import pytest
@pytest.mark.parametrize('test_input,expected', [
    pytest.param('sex', 0, marks=pytest.mark.basic),
    pytest.param('ethnicity', 0, marks=pytest.mark.basic),
    pytest.param('dcms_ageband', 0, marks=pytest.mark.basic),
    pytest.param('qualification', 0, marks=pytest.mark.xfail), # publication numbers dont add up - go through with penny - turn's out there is an extra column which is hidden by the publication called don't know which explains all this
    pytest.param('ftpt', 0, marks=pytest.mark.basic),
    pytest.param('nssec', 0, marks=pytest.mark.basic),
    pytest.param('region', 0, marks=pytest.mark.basic),
    pytest.param('cs', 0, marks=pytest.mark.basic),
    pytest.param('ci', 0, marks=pytest.mark.basic),
    pytest.param('culture', 0, marks=pytest.mark.basic),
    pytest.param('digital', 0, marks=pytest.mark.basic),
    pytest.param('gambling', 0, marks=pytest.mark.basic),
    pytest.param('sport', 0, marks=pytest.mark.basic),
    pytest.param('telecoms', 0, marks=pytest.mark.basic),
])
def test_datamatches(test_input, expected):
    assert sum((differencelist[test_input] < -0.05).any()) == expected
    assert sum((differencelist[test_input] > 0.05).any()) == expected


wb.save('dcms-testing2.xlsx')

# get final table with hierarchical indexes which I can check against those read in from excel (including order of rows etc), but then just output the values to the formatted excel templates

# for anonymisation, it seems something quite simple will work initially. Only gender, age, and nnsec seem like they will require rules

# for region the total won't = the sum anyway, so don't need to do annonymisation

# use openpyxl initially and move on to xlwings if necessary. xlsxwriter cannot read workbooks but should be considered if producing workbooks from scratch.













































